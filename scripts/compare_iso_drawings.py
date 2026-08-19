# joint_master(ISO Drawing DB)와 drawing.dwg_latest(ipcs-drawing 서버) 간 ISO Drawing 매칭 비교
import os
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill
from supabase import create_client, ClientOptions


def load_env():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    env_path = os.path.join(base_dir, ".env")
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ[k.strip()] = v.strip()


def fetch_all(sb, table, cols, page_size=1000):
    rows = []
    offset = 0
    while True:
        r = sb.table(table).select(cols).range(offset, offset + page_size - 1).execute()
        batch = r.data or []
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return rows


def main():
    load_env()

    ctrl_opts = ClientOptions(schema="construction", postgrest_client_timeout=90)
    ctrl_sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"], options=ctrl_opts)

    draw_opts = ClientOptions(schema="drawing", postgrest_client_timeout=120)
    draw_sb = create_client(os.environ["DRAWING_SUPABASE_URL"], os.environ["DRAWING_SUPABASE_KEY"], options=draw_opts)

    print("Fetching joint_master (ISO Drawing DB)...")
    jm_rows = fetch_all(ctrl_sb, "joint_master", "iso_drawing,rev")
    print(f"  {len(jm_rows)} rows")

    print("Fetching drawing.dwg_latest (ipcs-drawing server)...")
    dwg_rows = fetch_all(draw_sb, "dwg_latest", "drawing_no,revision")
    print(f"  {len(dwg_rows)} rows")

    # joint_master: iso_drawing 별 rev 값 최빈값(mode) 채택 (여러 joint가 같은 ISO를 공유)
    jm_rev_counter = {}
    for row in jm_rows:
        iso = (row.get("iso_drawing") or "").strip()
        if not iso:
            continue
        rev = (row.get("rev") or "").strip() or None
        jm_rev_counter.setdefault(iso, Counter())[rev] += 1

    jm_isos = set(jm_rev_counter.keys())
    jm_rev = {}
    for iso, counter in jm_rev_counter.items():
        counter.pop(None, None)
        jm_rev[iso] = counter.most_common(1)[0][0] if counter else None

    dwg_isos = set()
    dwg_rev = {}
    for row in dwg_rows:
        no = (row.get("drawing_no") or "").strip()
        if not no:
            continue
        dwg_isos.add(no)
        dwg_rev[no] = (row.get("revision") or "").strip() or None

    only_in_control = sorted(jm_isos - dwg_isos)
    only_in_drawing = sorted(dwg_isos - jm_isos)

    print(f"ISO Drawing DB에만 존재 (ipcs-drawing 서버에 없음): {len(only_in_control)}")
    print(f"ipcs-drawing 서버에만 존재 (ISO Drawing DB에 없음): {len(only_in_drawing)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unmatched ISO Drawings"

    headers = ["ISO Drawing", "구분", "Revision (ISO Drawing DB 기준)"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="305496")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for iso in only_in_control:
        ws.append([iso, "ISO Drawing DB에만 존재 (ipcs-drawing 서버 누락)", jm_rev.get(iso) or ""])

    for iso in only_in_drawing:
        # ipcs-drawing 서버에만 있는 항목은 ISO Drawing DB(joint_master)에 데이터 자체가 없음
        ws.append([iso, "ipcs-drawing 서버에만 존재 (ISO Drawing DB 누락)", ""])

    widths = [42, 40, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    out_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Reports")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "ISO_Drawing_Mismatch.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Total unmatched rows: {len(only_in_control) + len(only_in_drawing)}")


if __name__ == "__main__":
    main()
